#!/usr/bin/env python3
"""
全メソドロジー抽出スクリプト
- VROD Excel ファイルから全メソドロジー + プロジェクト数 + クレジット数を抽出
- CAD Trust API から全ユニークメソドロジーコードを収集
- 統合して src/data/all-methodologies.json に保存
"""

import json
import time
import urllib.request
import urllib.parse
import collections
import sys
import os

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ============================================================
# 定数
# ============================================================

VROD_FILE = os.path.join(os.path.dirname(__file__), "..", "VROD-registry-files--2025-12.xlsx")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "..", "src", "data", "all-methodologies.json")
CAD_TRUST_API = "https://observer.climateactiondata.org/api/v1/projects"
CAD_PAGE_LIMIT = 200


# ============================================================
# ヘルパー
# ============================================================

def fetch_json(url: str) -> dict:
    req = urllib.request.Request(
        url,
        headers={"Accept": "application/json", "User-Agent": "CarbonIntelligenceBot/1.0"}
    )
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode("utf-8"))


def infer_registry_from_code(name: str) -> str:
    """メソドロジーコード/名称から発行機関を推定（VROD Excelデータ用）"""
    n = name.strip()
    if not n:
        return "Unknown"
    # CDM codes
    if any(n.startswith(p) for p in ["ACM", "AMS-", "AR-ACM", "AR-AMS", "AR-AM", "AM0", "AM1", "AM2", "AM3", "AM4", "AM5", "AM6", "AM7", "AM8", "AM9"]):
        return "CDM"
    # Verra
    if n.startswith("VM") or n.startswith("VMR"):
        return "Verra"
    # ARB named protocols
    if n.startswith("ARB ") or "Compliance Offset Protocol" in n:
        return "ARB"
    # CAR protocols (common ones)
    if any(kw in n for kw in ["Improved Forest Management", "IFM", "Anti-Idling", "Plugging Orphan", "Destroying Methane from Livestock", "Nitrogen Management", "Rice Cultivation", "Urban Forest", "Grassland", "Organic Waste", "Landfill Gas"]):
        return "CAR"
    # ACR protocols
    if any(kw in n for kw in ["Advanced Refrigeration", "Certified Reclaimed HFC", "Transition to Advanced Formulation", "Destruction of Ozone Depleting", "High-Global Warming"]):
        return "ACR"
    return "Unknown"


def infer_registry_from_cad(raw: str) -> tuple[str, str]:
    """CAD Trust の 'Registry - MethodologyCode' 形式をパース → (registry, code)"""
    if not raw:
        return ("Unknown", raw)
    raw = raw.strip()
    sep = " - "
    if sep in raw:
        parts = raw.split(sep, 1)
        reg = parts[0].strip()
        code = parts[1].strip()
        # 正規化
        reg_map = {
            "CDM": "CDM",
            "VCS": "Verra",
            "GS": "Gold Standard",
            "GS4GG": "Gold Standard",
            "ACR": "ACR",
            "CAR": "CAR",
            "ART": "ART",
            "ARB": "ARB",
            "PURO": "Puro.earth",
            "JCM": "J-Credit",
        }
        normalized_reg = reg_map.get(reg.upper(), reg)
        return (normalized_reg, code)
    return ("Unknown", raw)


# ============================================================
# VROD Excel から抽出
# ============================================================

def extract_vrod_methodologies() -> dict[str, dict]:
    """
    VROD Excel の各 Projects シートから全メソドロジーを抽出
    Returns: {methodology_key: {name, registry, projects, credits, sources}}
    """
    print("=== VROD Excel からメソドロジー抽出中 ===")
    wb = openpyxl.load_workbook(VROD_FILE, read_only=True, data_only=True)

    result: dict[str, dict] = {}  # key = normalized name

    def add(name: str, registry: str, credits_val: int = 0):
        if not name or name.lower() in ("", "n/a", "none", "not applicable", "various", "multiple"):
            return
        name = name.strip()
        key = name.lower()
        if key not in result:
            result[key] = {
                "name": name,
                "registry": registry,
                "projects": 0,
                "credits": 0,
                "source": "vrod",
                "sources": [],
            }
        result[key]["projects"] += 1
        result[key]["credits"] += credits_val

    # --- ACR Projects ---
    ws = wb["ACR Projects"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    meth_col = next((i for i, h in enumerate(headers) if "Methodology" in h or "Protocol" in h), None)
    credits_col = next((i for i, h in enumerate(headers) if "Credits Registered" in h), None)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if meth_col is not None and row[meth_col]:
            name = str(row[meth_col]).strip()
            cred = 0
            if credits_col is not None and row[credits_col]:
                try:
                    cred = int(float(str(row[credits_col]).replace(",", "")))
                except:
                    pass
            add(name, "ACR", cred)
            count += 1
    print(f"  ACR Projects: {count} rows processed")

    # --- CAR Projects (Project Type をメソドロジーとして扱う) ---
    ws = wb["CAR Projects"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    type_col = next((i for i, h in enumerate(headers) if h == "Project Type"), None)
    credits_col = next((i for i, h in enumerate(headers) if "Credits Registered" in h), None)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if type_col is not None and row[type_col]:
            name = str(row[type_col]).strip()
            cred = 0
            if credits_col is not None and row[credits_col]:
                try:
                    cred = int(float(str(row[credits_col]).replace(",", "")))
                except:
                    pass
            add(name, "CAR", cred)
            count += 1
    print(f"  CAR Projects: {count} rows processed")

    # --- Gold Projects ---
    ws = wb["Gold Projects"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    meth_col = next((i for i, h in enumerate(headers) if h == "Methodology"), None)
    est_col = next((i for i, h in enumerate(headers) if "Estimated Annual" in h), None)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if meth_col is not None and row[meth_col]:
            name = str(row[meth_col]).strip()
            cred = 0
            if est_col is not None and row[est_col]:
                try:
                    cred = int(float(str(row[est_col]).replace(",", "")))
                except:
                    pass
            add(name, "Gold Standard", cred)
            count += 1
    print(f"  Gold Projects: {count} rows processed")

    # --- Verra Projects ---
    ws = wb["Verra Projects"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    meth_col = next((i for i, h in enumerate(headers) if h == "Methodology"), None)
    est_col = next((i for i, h in enumerate(headers) if "Estimated Annual" in h), None)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if meth_col is not None and row[meth_col]:
            # Verraは複数メソドロジーがセミコロン区切りの場合あり
            raw = str(row[meth_col]).strip()
            cred = 0
            if est_col is not None and row[est_col]:
                try:
                    cred = int(float(str(row[est_col]).replace(",", "")))
                except:
                    pass
            # セミコロン区切りを分割
            for name in raw.split(";"):
                name = name.strip()
                if name:
                    add(name, "Verra", cred)
            count += 1
    print(f"  Verra Projects: {count} rows processed")

    # --- ARB: Project Type から抽出 ---
    ws = wb["ARB Issuances & Retirements"]
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    type_col = next((i for i, h in enumerate(headers) if h == "Project Type"), None)
    credits_col = next((i for i, h in enumerate(headers) if "ARB Offset Credits Issued" in h), None)
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if type_col is not None and row[type_col]:
            name = str(row[type_col]).strip()
            cred = 0
            if credits_col is not None and row[credits_col]:
                try:
                    cred = int(float(str(row[credits_col]).replace(",", "")))
                except:
                    pass
            add(name, "ARB", cred)
            count += 1
    print(f"  ARB: {count} rows processed")

    wb.close()

    # credits を credits_total に集計（プロジェクト別合計として扱う）
    print(f"  VROD total unique methodologies: {len(result)}")
    return result


# ============================================================
# CAD Trust API から抽出
# ============================================================

def extract_cad_methodologies() -> dict[str, dict]:
    """
    CAD Trust API の全プロジェクトから methodology / methodology2 フィールドを収集
    Returns: {normalized_code: {name, registry, code, projects, source}}
    """
    print("\n=== CAD Trust API からメソドロジー抽出中 ===")

    # 最初に総ページ数確認
    first = fetch_json(f"{CAD_TRUST_API}?page=1&limit={CAD_PAGE_LIMIT}")
    total_pages = first.get("pageCount", 1)
    print(f"  総ページ数: {total_pages} (limit={CAD_PAGE_LIMIT})")

    counter: dict[str, dict] = {}  # key = raw methodology string

    def record(raw: str):
        if not raw:
            return
        raw = raw.strip()
        if raw not in counter:
            registry, code = infer_registry_from_cad(raw)
            counter[raw] = {
                "name": raw,
                "registry": registry,
                "code": code,
                "projects": 0,
                "source": "cad-trust",
            }
        counter[raw]["projects"] += 1

    # page 1 already fetched
    for proj in first.get("data", []):
        record(proj.get("methodology", ""))
        record(proj.get("methodology2", ""))

    for page in range(2, total_pages + 1):
        if page % 10 == 0:
            print(f"  ... page {page}/{total_pages}")
        try:
            data = fetch_json(f"{CAD_TRUST_API}?page={page}&limit={CAD_PAGE_LIMIT}")
            for proj in data.get("data", []):
                record(proj.get("methodology", ""))
                record(proj.get("methodology2", ""))
        except Exception as e:
            print(f"  Warning: page {page} failed: {e}")
        time.sleep(0.1)  # rate limit

    # None/empty は除外
    counter.pop("", None)
    counter.pop("None", None)

    print(f"  CAD Trust total unique methodology values: {len(counter)}")
    return counter


# ============================================================
# マージ & 出力
# ============================================================

def merge_and_save(vrod: dict, cad: dict):
    """VRODとCAD Trustのメソドロジーをマージして出力"""
    print("\n=== マージ中 ===")

    # VROD側でコードベースのメソドロジーを正規化
    # CAD Trustの "CDM - ACM0002" → code="ACM0002" を VROD の "ACM0002" と突合
    merged: dict[str, dict] = {}

    # VROD エントリを追加
    for key, v in vrod.items():
        merged[key] = {
            "name": v["name"],
            "registry": v["registry"],
            "projects_vrod": v["projects"],
            "credits_vrod": v["credits"],
            "projects_cad": 0,
            "source": ["vrod"],
        }

    # CAD Trust エントリをマージ
    for raw, c in cad.items():
        code = c["code"].lower()
        registry = c["registry"]
        cad_projects = c["projects"]
        raw_lower = raw.lower()
        code_lower = code.lower()

        # すでにVRODにあれば projects_cad に加算
        if code_lower in merged:
            merged[code_lower]["projects_cad"] += cad_projects
            if "cad-trust" not in merged[code_lower]["source"]:
                merged[code_lower]["source"].append("cad-trust")
        elif raw_lower in merged:
            merged[raw_lower]["projects_cad"] += cad_projects
            if "cad-trust" not in merged[raw_lower]["source"]:
                merged[raw_lower]["source"].append("cad-trust")
        else:
            # CAD Trust のみに存在
            merged[raw_lower] = {
                "name": c["code"] if c["code"] != raw else raw,
                "registry": registry,
                "projects_vrod": 0,
                "credits_vrod": 0,
                "projects_cad": cad_projects,
                "source": ["cad-trust"],
            }

    # 最終配列を生成（projects合計でソート）
    output = []
    for key, m in merged.items():
        total_projects = m["projects_vrod"] + m["projects_cad"]
        output.append({
            "name": m["name"],
            "registry": m["registry"],
            "projectsVrod": m["projects_vrod"],
            "creditsVrod": m["credits_vrod"],
            "projectsCad": m["projects_cad"],
            "totalProjects": total_projects,
            "source": m["source"],
        })

    # プロジェクト数降順でソート
    output.sort(key=lambda x: x["totalProjects"], reverse=True)

    # 統計サマリー
    summary = {
        "total": len(output),
        "vrodOnly": sum(1 for x in output if x["source"] == ["vrod"]),
        "cadOnly": sum(1 for x in output if x["source"] == ["cad-trust"]),
        "both": sum(1 for x in output if len(x["source"]) > 1),
        "byRegistry": {},
    }
    for item in output:
        reg = item["registry"]
        summary["byRegistry"][reg] = summary["byRegistry"].get(reg, 0) + 1

    result = {
        "generatedAt": __import__("datetime").datetime.now().isoformat(),
        "summary": summary,
        "methodologies": output,
    }

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n=== 完了 ===")
    print(f"総メソドロジー数: {len(output)}")
    print(f"  VROD のみ: {summary['vrodOnly']}")
    print(f"  CAD Trust のみ: {summary['cadOnly']}")
    print(f"  両方に存在: {summary['both']}")
    print(f"レジストリ別:")
    for reg, cnt in sorted(summary["byRegistry"].items(), key=lambda x: -x[1]):
        print(f"  {reg}: {cnt}件")
    print(f"\n出力先: {OUTPUT_FILE}")


# ============================================================
# メイン
# ============================================================

if __name__ == "__main__":
    vrod_data = extract_vrod_methodologies()
    cad_data = extract_cad_methodologies()
    merge_and_save(vrod_data, cad_data)
